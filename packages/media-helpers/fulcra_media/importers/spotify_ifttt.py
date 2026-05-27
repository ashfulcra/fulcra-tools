"""Spotify IFTTT->GDrive xlsx importer.

Years ago, users (including ash@) wired IFTTT applets that poll Spotify's
/me/player/recently-played and append rows to a Google Sheets spreadsheet.
The two common applets are 'Recent tracks' and 'Spotify Tracks V2', which
emit nearly identical 5-column rows:

  (timestamp_str, track_name, artist, spotify_track_id, spotify_url)

There's no header row. IFTTT renders the timestamp in the user's IFTTT
account timezone ('November 4, 2022 at 03:53PM'), not UTC.

Because two applets logged in parallel and each xlsx caps at ~2000 rows,
a long-running setup produces multiple overlapping files inside one zip:

  Spotify/Recent tracks.xlsx
  Spotify/Recent tracks (1).xlsx
  Spotify/Recent tracks (2).xlsx
  Spotify/Spotify Tracks V2.xlsx
  Spotify/Spotify Tracks V2 (1).xlsx
  ...

Cross-applet dedup key is (track_id, timestamp). Same track at different
times = real replay, keep both.
"""
from __future__ import annotations

import hashlib
import re
import zipfile
from collections.abc import Iterator
from datetime import datetime, timedelta, timezone, tzinfo
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from fulcra_common.cross_source_fingerprint import listened_fingerprint

from .base import NormalizedEvent, content_fingerprint

# "November 4, 2022 at 03:53PM"
_IFTTT_TS_RE = re.compile(
    r"^(?P<month>\w+)\s+(?P<day>\d{1,2}),\s+(?P<year>\d{4})\s+at\s+"
    r"(?P<hour>\d{1,2}):(?P<minute>\d{2})(?P<ampm>AM|PM)$"
)

_MONTHS = {
    "January": 1, "February": 2, "March": 3, "April": 4, "May": 5, "June": 6,
    "July": 7, "August": 8, "September": 9, "October": 10, "November": 11, "December": 12,
}


def parse_ifttt_timestamp(value: str, *, tz: tzinfo) -> datetime:
    """Parse 'November 4, 2022 at 03:53PM' into a tz-aware datetime in `tz`."""
    m = _IFTTT_TS_RE.match(value or "")
    if not m:
        raise ValueError(f"not an IFTTT timestamp: {value!r}")
    month = _MONTHS.get(m.group("month"))
    if month is None:
        raise ValueError(f"unknown month in IFTTT timestamp: {value!r}")
    hour = int(m.group("hour")) % 12
    if m.group("ampm") == "PM":
        hour += 12
    return datetime(
        int(m.group("year")), month, int(m.group("day")),
        hour, int(m.group("minute")),
        tzinfo=tz,
    )


def _coerce_str(v: object) -> str:
    """Spreadsheets coerce numeric-looking track names to floats — round-trip cleanly."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _det_id(track_id: str, ts: datetime) -> str:
    h = hashlib.sha256(f"{track_id}|{ts.isoformat()}".encode()).hexdigest()
    return f"com.fulcra.media.spotify-ifttt.v1.{h[:16]}"


def _iter_xlsx_rows(zf: zipfile.ZipFile) -> Iterator[tuple]:
    """Yield rows from every .xlsx member inside the zip (Spotify/ subdir)."""
    for name in zf.namelist():
        if not name.lower().endswith(".xlsx"):
            continue
        with zf.open(name) as f:
            data = f.read()
        wb = load_workbook(BytesIO(data), read_only=True)
        try:
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                if row and row[0]:
                    yield row
        finally:
            wb.close()


def parse_ifttt_zip(
    zip_path: Path,
    *,
    tz: tzinfo | None = None,
) -> Iterator[NormalizedEvent]:
    """Parse a Spotify IFTTT->GDrive zip into NormalizedEvents.

    `tz`: the timezone IFTTT rendered the timestamps in. Defaults to UTC.
    Users whose IFTTT account was in a local zone should pass that zone
    (e.g. ZoneInfo("America/New_York")) for correct play times.
    """
    tz = tz or timezone.utc
    seen: dict[tuple[str, datetime], NormalizedEvent] = {}

    with zipfile.ZipFile(zip_path) as zf:
        for row in _iter_xlsx_rows(zf):
            ts_raw, track_raw, artist_raw, track_id, url = (row + (None,) * 5)[:5]
            track_id = _coerce_str(track_id)
            if not track_id:
                continue
            ts = parse_ifttt_timestamp(_coerce_str(ts_raw), tz=tz)
            key = (track_id, ts)
            if key in seen:
                continue
            artist = _coerce_str(artist_raw)
            track = _coerce_str(track_raw)
            if not artist or not track:
                continue
            note = f"{artist} – {track}"
            fp = content_fingerprint("music", artist=artist, track=track)
            cross = listened_fingerprint(timestamp=ts, artist=artist, track=track)
            seen[key] = NormalizedEvent(
                importer="spotify-ifttt",
                service="spotify",
                category="listened",
                note=note,
                title=track,
                start_time=ts,
                end_time=ts + timedelta(seconds=1),
                deterministic_id=_det_id(track_id, ts),
                timestamp_confidence="medium",
                external_ids={
                    "track_id": track_id,
                    "artist": artist,
                    "track": track,
                    "spotify_url": _coerce_str(url),
                    "content_fingerprint": fp,
                    "point_in_time": True,
                    "source_applet": "ifttt-gdrive",
                },
                extra_source_ids=(cross,) if cross else (),
            )

    for _, event in sorted(seen.items(), key=lambda kv: kv[1].start_time):
        yield event
